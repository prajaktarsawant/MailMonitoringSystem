import win32com.client
import time
import threading
from winotify import Notification
import tkinter as tk
from tkinter import simpledialog
import pythoncom
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pyttsx3
import pandas as pd 
from bs4 import BeautifulSoup 
import re
from io import StringIO

# Define the EmailRecord class
class EmailRecord:
    def __init__(self, no, folder_name, email_sender, email_receiver, subject, timestamp, df, background_color): #, body
        self.No = no
        self.FolderName = folder_name
        self.EmailSender = email_sender
        self.EmailReceiver = email_receiver
        self.Subject = subject
        self.TimeStamp = timestamp
        self.df = df 
        self.background_color = background_color
        #self.Body = body

def add_records_to_excel(records):
    if records:
        # Get today's date in YYYY-MM-DD format
        today = datetime.now().strftime("%Y-%m-%d")
        # Path to the reports folder
        folder_path = './reports'
        # Ensure the folder exists
        os.makedirs(folder_path, exist_ok=True)
        # Path to the Excel file
        file_path = os.path.join(folder_path, f'{today}.xlsx')

        # Check if the file exists
        if os.path.isfile(file_path):
            # Load the existing workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # Find the current maximum "No" in the sheet
            max_no = max((sheet.cell(row=i, column=1).value or 0) for i in range(2, sheet.max_row + 1))
        else:
            # Create a new workbook and sheet
            workbook = Workbook()
            sheet = workbook.active
            # Define the header
            header = ['No', 'FolderName', 'EmailSender', 'EmailReceiver', 'Subject', 'TimeStamp']#, 'Body'
            sheet.append(header)
            max_no = 0  # No records initially

        # Write the records to the Excel file
        for index, record in enumerate(records, start=max_no + 1):
            # Remove timezone information from the timestamp
            timestamp_without_tz = record.TimeStamp#.str.replace(tzinfo=None)
            sheet.append([
                index,
                record.FolderName,
                record.EmailSender,
                record.EmailReceiver,
                record.Subject,
                str(timestamp_without_tz),
                #record.Body
            ])

        # Save the workbook
        workbook.save(file_path)

        # Return the last record
        if records:
            return records[-1]
    return []

def get_folder_by_name(namespace, folder_name, parent_folder=None):
    """Recursively searches for a folder by name."""
    if parent_folder is None:
        parent_folder = namespace.Folders.Item(1)

    for folder in parent_folder.Folders:
        if folder.Name == folder_name:
            return folder
        else:
            subfolder = get_folder_by_name(namespace, folder_name, folder)
            if subfolder:
                return subfolder
    return None

#Monitor Outlook folder, notify new mails and store data
def monitor_outlook_folder(folder_name, text_widget, stop_time):
    try:
        text_widget.insert(tk.END, f"\nEnd Time '{stop_time}'...\n")
        # Initialize COM for Outlook
        pythoncom.CoInitialize()

        # Initialize Outlook application
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        
        # Get the folder
        folder = get_folder_by_name(outlook, folder_name)

        if not folder:
            text_widget.insert(tk.END, f"Folder '{folder_name}' not found.\n")
            return

        # Fetch messages from the folder
        messages = folder.Items
        last_email_time = None
        email_count = count_all_emails(folder_name)

        text_widget.insert(tk.END, f"\nMonitoring folder '{folder_name}'...\n")
        while True:
            # Exit loop if current time is past the stop_time
            if datetime.now().time() > datetime.strptime(stop_time, "%H:%M").time():
                text_widget.insert(tk.END, f"Stop time reached, stopping monitoring of folder '{folder_name}'.\n")
                text_widget.yview(tk.END)
                break  # Exit the while loop

            current_email = messages.GetLast()
            if current_email:
                current_email_time = current_email.ReceivedTime
                if current_email_time != last_email_time:
                    new_new_counts = count_all_emails(folder_name) - email_count
                    email_count = count_all_emails(folder_name)
                    new_emails = get_last_n_emails(folder_name, new_new_counts)
                    last_email_time = current_email_time
                    subject = current_email.Subject
                    sender = current_email.SenderName
                    if folder_name in ['Ariston']:
                        for i in range(2):
                            text_to_speech(f"{folder_name} Alert detected, please check")
                        for email in new_emails:
                            try:
                                get_table_data_to_excel(email.df, email.background_color)
                            except Exception as e:
                                text_widget.insert(tk.END, f"error for 'get_table_data_to_excel' as:{e}")
                                
                    if folder_name == 'Canon' and sender == 'alert@appdynamics.com':
                        file = 'Canon.xlsx'
                        for i in range(2):
                            text_to_speech(f"{folder_name} Alert detected, please check")

                        columns = ['Sr. No.', 'Subject', 'DateTime', 'Month']                            
                        create_excel_file_if_not_exists(file, columns)

                        # Load existing data to find the last Sr. No.
                        existing_data = pd.read_excel(file)
                        last_sr_no = existing_data['Sr. No.'].max() if not existing_data.empty else 0  
                        new_records = []
                        for e in new_emails:
                            last_sr_no += 1  # Increment Sr. No.
                            e_tuple = (last_sr_no, e.Subject, datetime.now().strftime('%a %d-%m-%Y %I:%M %p'), datetime.now().strftime('%b-%y'))
                            new_records.append(e_tuple)   

                        # Add new records to Excel
                        for record in new_records:
                            add_record_to_excel(filepath=file, record=record)

                    #Personal mail detection for testing
                    if folder_name == 'Personal':
                        file = 'Personal.xlsx'
                        for i in range(2):
                            text_to_speech(f"{folder_name} Alert detected, please check")

                        # Load existing data to find the last Sr. No.
                        columns = ['Sr. No.', 'Subject', 'DateTime', 'Month']
                        existing_data = pd.read_excel(file)
                        last_sr_no = existing_data['Sr. No.'].max() if not existing_data.empty else 0  
                        new_records = []
                        
                        for e in new_emails:
                            create_excel_file_if_not_exists(file, columns)
                            last_sr_no += 1  # Increment Sr. No.
                            e_tuple = (last_sr_no, e.Subject, datetime.now().strftime('%a %d-%m-%Y %I:%M %p'), datetime.now().strftime('%b-%y'))                            
                            new_records.append(e_tuple)   

                        # Add new records to Excel
                        for record in new_records:
                            add_record_to_excel(filepath=file, record=record)


                    # Try to add records to the Excel file
                    try:
                        record = add_records_to_excel(new_emails)
                    except Exception as excel_error:
                        text_widget.insert(tk.END, f"Error updating Excel file: {excel_error}\n")
                        text_widget.yview(tk.END)
                        # Attempt to recreate the file if corrupted
                        continue
                                    

                    # Notify user and start TTS
                    status_message = f"\nNew Email ({folder_name}):\n(time: {current_email.ReceivedTime}) \n{subject} \nfrom {sender}\n"
                    text_widget.insert(tk.END, status_message)
                    text_widget.yview(tk.END)  # Scroll to the bottom

                    # Show desktop notification using winotify
                    toast = Notification(
                        app_id="Outlook Email Monitor",
                        title="New Email Notification",
                        msg=f"Subject: {subject}\nFrom: {sender}",
                    )
                    toast.show()
                    last_email_time = current_email_time
            else:
                # If no email is found, exit the loop and close the thread
                text_widget.insert(tk.END, "No new emails found, stopping monitoring.\n")
                text_widget.yview(tk.END)
                break  # Exit the while loop
            time.sleep(10)  # Check every 10 seconds

    except Exception as e:
        text_widget.insert(tk.END, f"An error occurred (monitor_outlook_folder): {e}\n")
        text_widget.yview(tk.END)


#Overall mail records for all folders which user is monitoring
def add_record_to_excel(filepath, record):
    workbook = load_workbook(filepath)
    sheet = workbook.active
    next_row = sheet.max_row + 1
    # Find the last "Sr. No." (assuming it's in column A)
    last_row = sheet.max_row
    last_sr_no = sheet.cell(row=last_row, column=1).value  # Read last "Sr. No."
    
    # If no Sr. No. exists yet, start from 1
    if not isinstance(last_sr_no, int):
        last_sr_no = 0  

    next_sr_no = last_sr_no + 1  # Increment Sr. No.

    # Insert the new record with the auto-incremented Sr. No.
    next_row = last_row + 1
    sheet.cell(row=next_row, column=1, value=next_sr_no)  # Sr. No.

    # Insert the rest of the data
    for col, value in enumerate(record, start=1):
        sheet.cell(row=next_row, column=col, value=value)
    workbook.save(filepath)


#Creating file if not exist
def create_excel_file_if_not_exists(filepath, columns):
    if not os.path.exists(filepath):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(columns)  # Add column headers
        workbook.save(filepath)


#Ariston Tasks and Tickets classification and data collection
def get_table_data_to_excel(df: pd.DataFrame, background_color: str):
    # Add timestamp and month columns
    df['DateTime'] = datetime.now().strftime('%a %d-%m-%Y %I:%M %p') # Ddd DD-MM-YYYY HH:MM AM/PM
    df['Month'] = datetime.now().strftime('%b-%y')
    columns_of_interest = ['Ticket number', 'Priority', 'DateTime', 'Month']
    df_filtered = df[columns_of_interest]
    if background_color == '#dbdb07':  # For Tickets
        ticket_file = 'Ticket.xlsx'
        create_excel_file_if_not_exists(ticket_file, columns_of_interest)

        # Load existing tickets from the file
        existing_tickets = pd.read_excel(ticket_file)['Ticket number'].tolist()
        count_of_existing_records = len(existing_tickets)

        # Filter out duplicates (only new tickets)
        new_tickets = df_filtered[~df_filtered['Ticket number'].isin(existing_tickets)]
        # new_tickets.to_csv("new ticket got on line number 269.csv")

        # Append only new records to the file
        if not new_tickets.empty:
            num_new_records = new_tickets.shape[0]  # Get row count    
            # Generate new "Sr. No." column
            new_tickets.insert(0, "Sr. No.", [count_of_existing_records + i for i in range(1, num_new_records + 1)])
            with pd.ExcelWriter(ticket_file, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                new_tickets.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

    elif background_color == '#04b894':  # For Tasks
        task_file = 'Task.xlsx'
        create_excel_file_if_not_exists(task_file, columns_of_interest)

        # Load existing tasks from the file
        existing_tasks = pd.read_excel(task_file)['Ticket number'].tolist()
        count_of_existing_records = len(existing_tasks)

        # Filter out duplicates (only new tasks)
        new_tasks = df_filtered[~df_filtered['Ticket number'].isin(existing_tasks)]

        # Append only new records to the file
        if not new_tasks.empty:
            num_new_records = new_tasks.shape[0]  # Get row count    
            # Generate new "Sr. No." column
            new_tasks.insert(0, "Sr. No.", [count_of_existing_records + i for i in range(1, num_new_records + 1)])
            with pd.ExcelWriter(task_file, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                new_tasks.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    else:
        pass



#Converting mail to HTML format to check any table is present or not
def parse_html_table_and_apply_style(html_content):
    # Parse the HTML using BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')
    background_color = extract_background_color_from_style(html_content)
    # Find the first table in the HTML content
    table = soup.find('table')

    if table:
        # Use pandas read_html to convert the table into a DataFrame
        df = pd.read_html(StringIO(str(table)))[0]
        return df, background_color #, styled_df
    else:
        print("No table found in the email.")
        return None

#Finding table heading color of Ariston
def extract_background_color_from_style(html_content):
    try:
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(html_content, "html.parser")
        
        # Check for the inline style in the <thead> tag
        thead = soup.find("thead")
        if thead:
            # First, check the inline style attribute for background color
            style = thead.get("style")
            if style and "background" in style:
                background_color = style.split("background:")[1].split(";")[0].strip()
                return background_color
        
        # If no inline style, check for background color in <style> tag
        style_tag = soup.find("style")
        if style_tag:
            # Parse the style tag for the background color
            style_content = style_tag.string
            if style_content:
                start_index = style_content.find("background:")
                if start_index != -1:
                    # Extract the color code from CSS content
                    color_code = style_content[start_index:].split("background:")[1].split(";")[0].strip()
                    return color_code

        print("No background color found in <thead> or <style>.")
        return None

    except Exception as e:
        print(f"Error extracting header color: {e}")
        return None

    except Exception as e:
        print(f"Error extracting header color: {e}")
        return None

def get_last_n_emails(folder_name, count_of_emails):
    try:
        # Initialize COM for Outlook
        pythoncom.CoInitialize()

        # Initialize Outlook application
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

        # Get the folder
        folder = get_folder_by_name(outlook, folder_name)

        if not folder:
            print(f"Folder '{folder_name}' not found.")
            return []

        # Fetch messages from the folder
        messages = folder.Items
        messages.Sort("[ReceivedTime]", True)  # Sort messages by received time in descending order

        # Retrieve the last 'count_of_emails' messages
        email_records = []
        for i in range(min(count_of_emails, len(messages))):
            message = messages[i]
            if folder_name == 'Ariston':
                df, background_color = parse_html_table_and_apply_style(message.HTMLBody)
                email_record = EmailRecord(
                    no=i + 1,
                    folder_name=folder_name,
                    email_sender=message.SenderName,
                    email_receiver=message.To,
                    subject=message.Subject,
                    timestamp=str(message.ReceivedTime),
                    df = df,
                    background_color=background_color
                    #body=message.Body
                )    
            else:
                email_record = EmailRecord(
                    no=i + 1,
                    folder_name=folder_name,
                    email_sender=message.SenderName,
                    email_receiver=message.To,
                    subject=message.Subject,
                    timestamp=message.ReceivedTime,
                    df = None,
                    background_color=None
                    #body=message.Body
                )  

            email_records.append(email_record)

        return email_records

    except Exception as e:
        print(f"An error occurred (get_last_n_emails): {e}")
        return []

def count_all_emails(folder_name):
    try:
        # Initialize COM for Outlook
        pythoncom.CoInitialize()

        # Initialize Outlook application
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

        # Get the folder
        folder = get_folder_by_name(outlook, folder_name)

        if not folder:
            print(f"Folder '{folder_name}' not found.")
            return 0

        # Fetch messages from the folder
        messages = folder.Items

        # Return the count of messages
        return len(messages)

    except Exception as e:
        print(f"An error occurred (count_all_emails): {e}")
        return 0    


def text_to_speech(text):
    # Initialize the TTS engine
    engine = pyttsx3.init()
    
    # Set properties (optional)
    # Example: Setting the speech rate
    rate = engine.getProperty('rate')   # Getting current speaking rate
    engine.setProperty('rate', rate-50) # Setting a new speaking rate

    # Example: Setting the voice (male/female)
    voices = engine.getProperty('voices')
    engine.setProperty('voice', voices[0].id)  # 0 for male, 1 for female

    # Convert text to speech
    engine.say(text)

    # Wait for the speaking to finish
    engine.runAndWait()

#To start moitoring
def start_monitoring(folders, text_widget,stop_time):
    threads = []
    for folder_name in folders:
        thread = threading.Thread(target=monitor_outlook_folder, args=(folder_name, text_widget,stop_time))
        thread.start()
        threads.append(thread)

#Auto-stop monitoring at the time defined by the user
def stop_monitoring_at(end_time, text_widget):
        # Parse the end time
        now = datetime.now()
        end_time_dt = datetime.strptime(end_time, "%H:%M").replace(
            year=now.year, month=now.month, day=now.day
        )

        # Calculate the time remaining
        if end_time_dt < now:
            text_widget.insert(tk.END, "The specified end time is in the past. Monitoring will not start.\n")
            text_widget.yview(tk.END)
            return False, None
        return True, end_time_dt


def main():
    # Create the main window
    root = tk.Tk()
    root.title("Outlook Folder Monitor")

    # Create a Text widget to display status updates
    text_widget = tk.Text(root, height=20, width=120)
    text_widget.pack(padx=10, pady=10)
    # Variable to hold monitoring threads
    monitoring_threads = []

    # Create a button to start monitoring
    def start():
        folder_names = simpledialog.askstring("Folder Names", "Enter folder names to monitor:")
        if folder_names:
            folders = [name.strip() for name in folder_names.split(',')]
            stop_time = simpledialog.askstring("Stop Time", "Enter monitoring stop time in 24hr HH:MM format:")
            correct, end_time = stop_monitoring_at(stop_time,text_widget)
            if(correct):
                start_monitoring(folders, text_widget,stop_time)
                text_widget.insert(tk.END, "Monitoring started...\n")
                text_to_speech(f"Monitoring started for {folder_names} project")
                text_widget.yview(tk.END)            
        else:
            text_widget.insert(tk.END, "No folder names entered.\n")
            text_widget.yview(tk.END)

    start_button = tk.Button(root, text="Select Folder", command=start)
    start_button.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
